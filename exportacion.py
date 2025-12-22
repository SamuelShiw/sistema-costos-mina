# exportacion.py
import pandas as pd
import io
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- IMPORTANTE: El nombre de esta función debe ser EXACTO ---
def generar_excel_corporativo(df_detalle, df_resumen, usuario_descarga, rol_descarga):
    """
    Genera un Excel de Alto Nivel con formato ejecutivo, auditoría y diseño limpio.
    """
    
    output = io.BytesIO()
    
    # Definimos colores corporativos (Azul Acero para cabeceras)
    COLOR_CABECERA = "1F4E78" 
    COLOR_TEXTO_CABECERA = "FFFFFF"
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        
        # --- 1. PREPARACIÓN DE DATOS ---
        
        # Hoja Resumen (KPIs)
        df_res_final = df_resumen.rename(columns={
            'labor': 'Labor / Frente',
            'tipo_labor': 'Tipo de Labor',
            'avance': 'Avance (m)',
            'tm': 'Mineral Extraído (TM)',
            'costo_pen': 'Gasto Total (S/)'
        })
        
        # Cálculos extra
        if 'Avance (m)' in df_res_final.columns and 'Gasto Total (S/)' in df_res_final.columns:
            df_res_final['Costo Unitario (S/m)'] = df_res_final.apply(
                lambda x: x['Gasto Total (S/)'] / x['Avance (m)'] if x['Avance (m)'] > 0 else 0, axis=1
            )

        # Hoja Detalle (Base de Datos)
        cols_map = {
            'fecha': 'Fecha',
            'guardia': 'Guardia',
            'labor': 'Ubicación (Labor)',
            'tipo_labor': 'Tipo',
            'categoria': 'Rubro / Categoría',
            'insumo': 'Material Utilizado',
            'unidad': 'Unidad',
            'cantidad': 'Cant. Consumida',
            'precio_unit': 'Precio Unit. (S/)',
            'costo_pen': 'Costo Total (S/)',
            'avance': 'Avance del Disparo (m)',
            'usuario_registro': 'Digitado Por'
        }
        
        cols_existentes = [c for c in df_detalle.columns if c in cols_map.keys()]
        df_det_final = df_detalle[cols_existentes].rename(columns=cols_map)
        
        # --- 2. ESCRITURA ---
        df_res_final.to_excel(writer, index=False, sheet_name='Resumen Gerencial', startrow=4)
        df_det_final.to_excel(writer, index=False, sheet_name='Base de Datos Detallada', startrow=4)
        
        # --- 3. MAQUILLAJE PROFESIONAL ---
        workbook = writer.book
        
        header_font = Font(name='Calibri', size=11, bold=True, color=COLOR_TEXTO_CABECERA)
        header_fill = PatternFill(start_color=COLOR_CABECERA, end_color=COLOR_CABECERA, fill_type='solid')
        border_style = Side(border_style="thin", color="000000")
        border_box = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
        currency_fmt = '#,##0.00 "S/"'
        number_fmt = '#,##0.00'
        
        def formatear_hoja(sheet_name, titulo_reporte):
            ws = workbook[sheet_name]
            
            # Encabezado
            ws['A1'] = "REPORTE DE COSTOS OPERATIVOS - MINECOST"
            ws['A1'].font = Font(size=16, bold=True, color="1F4E78")
            ws['A2'] = f"Generado por: {usuario_descarga.upper()} ({rol_descarga})"
            ws['A2'].font = Font(size=10, italic=True, color="555555")
            ws['A3'] = f"Fecha de Emisión: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            ws['A3'].font = Font(size=10, italic=True, color="555555")
            
            # Tabla
            max_row = ws.max_row
            max_col = ws.max_column
            
            for col in range(1, max_col + 1):
                col_letter = get_column_letter(col)
                
                # Cabecera
                cell_header = ws[f"{col_letter}5"]
                cell_header.fill = header_fill
                cell_header.font = header_font
                cell_header.alignment = Alignment(horizontal='center', vertical='center')
                cell_header.border = border_box
                
                length = len(str(cell_header.value))
                ws.column_dimensions[col_letter].width = length + 5
                
                # Contenido
                for row in range(6, max_row + 1):
                    cell = ws[f"{col_letter}{row}"]
                    cell.border = border_box
                    
                    if isinstance(cell.value, (int, float)):
                        header_text = str(cell_header.value).lower()
                        if 's/' in header_text or 'costo' in header_text or 'precio' in header_text or 'gasto' in header_text:
                            cell.number_format = currency_fmt
                        else:
                            cell.number_format = number_fmt
                    else:
                        cell.alignment = Alignment(horizontal='left')

        formatear_hoja('Resumen Gerencial', "RESUMEN EJECUTIVO")
        formatear_hoja('Base de Datos Detallada', "DETALLE DE MOVIMIENTOS")

    return output.getvalue()
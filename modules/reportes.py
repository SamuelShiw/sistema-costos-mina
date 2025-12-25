# modules/reportes.py
import pandas as pd
import io
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference

def preparar_resumen(df_agrupado):
    """
    Prepara el dataframe de resumen (agrupado por labor) para el reporte.
    """
    cols_map = {
        'labor': 'Labor / Ubicación',
        'avance': 'Avance (m)',
        'mineral_tm': 'Mineral Extraído (TM)',
        'precio_total': 'Gasto Total (S/)'
    }
    cols_existentes = [c for c in df_agrupado.columns if c in cols_map.keys()]
    df_final = df_agrupado[cols_existentes].rename(columns=cols_map)

    if 'Avance (m)' in df_final.columns and 'Gasto Total (S/)' in df_final.columns:
        df_final['Costo Unitario (S/m)'] = df_final.apply(
            lambda x: x['Gasto Total (S/)'] / x['Avance (m)'] if x['Avance (m)'] > 0 else 0,
            axis=1
        )
    return df_final

def preparar_detalle(df_detalle):
    """
    Prepara el dataframe detallado (registro por registro).
    """
    cols_map = {
        'fecha': 'Fecha',
        'guardia': 'Guardia',
        'labor': 'Ubicación (Labor)',
        'categoria': 'Rubro / Categoría',
        'detalle': 'Material / Detalle',
        'unidad': 'Unidad',
        'cantidad': 'Cant. Consumida',
        'precio_total': 'Costo Total (S/)',
        'avance': 'Avance (m)',
        'mineral_tm': 'Mineral (TM)',
        'usuario_registro': 'Digitado Por'
    }
    cols_existentes = [c for c in df_detalle.columns if c in cols_map.keys()]
    return df_detalle[cols_existentes].rename(columns=cols_map)

def generar_excel_corporativo(df_detalle, df_resumen, usuario="Admin", rol="Lector"):
    """
    Función principal que genera el Excel binario en memoria.
    """
    output = io.BytesIO()

    COLOR_CABECERA = "1F4E78"
    COLOR_TEXTO_CABECERA = "FFFFFF"

    df_res_final = preparar_resumen(df_resumen)
    df_det_final = preparar_detalle(df_detalle)

    pivot_chart = pd.pivot_table(
        df_det_final,
        values='Costo Total (S/)',
        index='Rubro / Categoría',
        aggfunc='sum'
    ).reset_index().sort_values(by='Costo Total (S/)', ascending=False)

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_res_final.to_excel(writer, index=False, sheet_name='Resumen Gerencial', startrow=4)
        df_det_final.to_excel(writer, index=False, sheet_name='Base de Datos Detallada', startrow=4)
        pivot_chart.to_excel(writer, index=False, sheet_name='Análisis Gráfico', startrow=4)

        workbook = writer.book

        header_font = Font(name='Calibri', size=11, bold=True, color=COLOR_TEXTO_CABECERA)
        header_fill = PatternFill(start_color=COLOR_CABECERA, end_color=COLOR_CABECERA, fill_type='solid')
        border_style = Side(border_style="thin", color="000000")
        border_box = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
        currency_fmt = '#,##0.00 "S/"'
        number_fmt = '#,##0.00'

        def formatear_hoja(sheet_name, titulo):
            ws = workbook[sheet_name]
            ws['A1'] = f"REPORTES CORE - {titulo}"
            ws['A1'].font = Font(size=16, bold=True, color=COLOR_CABECERA)
            ws['A2'] = f"Generado por: {usuario} ({rol})"
            ws['A2'].font = Font(size=10, italic=True, color="555555")
            ws['A3'] = f"Fecha de Emisión: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
            ws['A3'].font = Font(size=10, italic=True, color="555555")

            max_row = ws.max_row
            max_col = ws.max_column

            for col in range(1, max_col + 1):
                col_letter = get_column_letter(col)
                cell_header = ws[f"{col_letter}5"]
                cell_header.fill = header_fill
                cell_header.font = header_font
                cell_header.alignment = Alignment(horizontal='center', vertical='center')
                cell_header.border = border_box
                ws.column_dimensions[col_letter].width = len(str(cell_header.value)) + 8

                for row in range(6, max_row + 1):
                    cell = ws[f"{col_letter}{row}"]
                    cell.border = border_box
                    if isinstance(cell.value, (int, float)):
                        header_text = str(cell_header.value).lower()
                        if 's/' in header_text or 'costo' in header_text or 'gasto' in header_text or 'precio' in header_text:
                            cell.number_format = currency_fmt
                        else:
                            cell.number_format = number_fmt
                    else:
                        cell.alignment = Alignment(horizontal='left')

            ws.auto_filter.ref = f"A5:{get_column_letter(max_col)}{max_row}"
            ws.freeze_panes = "A6"

        formatear_hoja('Resumen Gerencial', "RESUMEN EJECUTIVO DE COSTOS")
        formatear_hoja('Base de Datos Detallada', "REGISTRO DETALLADO DE CONSUMOS")
        formatear_hoja('Análisis Gráfico', "ANÁLISIS POR CATEGORÍA")

        ws_pivot = workbook['Análisis Gráfico']

        # Gráfico de Barras
        chart_bar = BarChart()
        chart_bar.type = "bar"
        chart_bar.style = 10
        chart_bar.title = "Gasto Total por Categoría (Pareto)"
        chart_bar.y_axis.title = "Soles (S/)"
        chart_bar.x_axis.title = "Categoría"
        data = Reference(ws_pivot, min_col=2, min_row=5, max_row=ws_pivot.max_row)
        cats = Reference(ws_pivot, min_col=1, min_row=6, max_row=ws_pivot.max_row)
        chart_bar.add_data(data, titles_from_data=True)
        chart_bar.set_categories(cats)
        ws_pivot.add_chart(chart_bar, "E5")

        # Gráfico Circular
        chart_pie = PieChart()
        chart_pie.title = "Distribución de Costos por Categoría"
        chart_pie.add_data(data, titles_from_data=True)
        chart_pie.set_categories(cats)
        ws_pivot.add_chart(chart_pie, "E20")

        # Gráfico de Columnas
        chart_col = BarChart()
        chart_col.type = "col"
        chart_col.style = 11
        chart_col.title = "Comparación de Costos por Categoría"
        chart_col.y_axis.title = "Soles (S/)"
        chart_col.x_axis.title = "Categoría"
        chart_col.add_data(data, titles_from_data=True)
        chart_col.set_categories(cats)
        ws_pivot.add_chart(chart_col, "M5")

        # Gráfico de Líneas
        chart_line = LineChart()
        chart_line.title = "Tendencia de Costos"
        chart_line.style = 13
        chart_line.y_axis.title = "Soles (S/)"
        chart_line.add_data(data, titles_from_data=True)
        chart_line.set_categories(cats)
        ws_pivot.add_chart(chart_line, "M20")

    output.seek(0)
    return output.getvalue()